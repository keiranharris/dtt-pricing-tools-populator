"""
Excel Constants Reader Module

This module provides atomic functions for reading constant data from Excel files
that are used to populate pricing tool spreadsheets. Handles graceful error
scenarios and supports both .xlsx and .xlsb file formats.

Dependencies:
- openpyxl: Excel file reading/writing
- pathlib: Path handling
- typing: Type hints

Author: DTT Pricing Tool Accelerator
Feature: 002-excel-data-population
"""

from pathlib import Path
from typing import Dict, Optional
import logging

# Import SpecKit data models  
from data_models import ConstantsData

# Configure logging
logger = logging.getLogger(__name__)


def read_constants_data(constants_dir: Path, filename: str) -> Dict[str, str]:
    """
    Read constants data from Excel file mapping Column C (field names) to Column D (values).
    
    Args:
        constants_dir: Directory containing the constants file
        filename: Name of the constants Excel file
        
    Returns:
        Dictionary mapping field names to values, empty dict if file missing
        
    Example:
        >>> constants = read_constants_data(Path("00-CONSTANTS"), "constants.xlsx")
        >>> print(constants.get("Opportunity ID", "Not found"))
    """
    file_path = constants_dir / filename
    
    # Graceful handling for missing files
    if not file_path.exists():
        logger.info(f"Constants file not found: {file_path}")
        return {}
    
    try:
        # Import openpyxl and constants
        import openpyxl
        from constants import PRICING_SETUP_CONSTANTS_FIELD_COL, PRICING_SETUP_CONSTANTS_VALUE_COL

        # Load workbook (supports both .xlsx and .xlsb)
        workbook = openpyxl.load_workbook(file_path, data_only=True)

        # Validate "Pricing Setup" worksheet exists
        if "Pricing Setup" not in workbook.sheetnames:
            logger.warning(f"'Pricing Setup' worksheet not found in {filename}")
            return {}

        worksheet = workbook["Pricing Setup"]

        # Use hard-coded columns for field names/values
        field_mapping = parse_field_value_mapping(
            worksheet,
            PRICING_SETUP_CONSTANTS_FIELD_COL,
            PRICING_SETUP_CONSTANTS_VALUE_COL
        )

        logger.info(f"Successfully loaded {len(field_mapping)} constants from {filename}")
        return field_mapping

    except ImportError:
        logger.error("openpyxl library not available - cannot read Excel files")
        return {}
    except Exception as e:
        logger.error(f"Error reading constants file {file_path}: {e}")
        return {}


def read_constants_data_structured(constants_file: Path) -> ConstantsData:
    """
    Read field mappings from constants Excel file.
    
    Args:
        constants_file: Path to constants Excel file
        
    Returns:
        ConstantsData: Loaded field mappings and metadata
        
    Raises:
        FileNotFoundError: If constants file doesn't exist
        ValueError: If file format is invalid
        PermissionError: If file is not readable
    """
    if not constants_file.exists():
        raise FileNotFoundError(f"Constants file not found: {constants_file}")
    
    if not constants_file.is_file():
        raise ValueError(f"Constants path is not a file: {constants_file}")
    
    try:
        # Use existing read_constants_data function
        constants_dict = read_constants_data(constants_file.parent, constants_file.name)
        
        # Create ConstantsData model
        return ConstantsData(
            constants_file_path=constants_file,
            constants_fields=constants_dict,
            file_size_bytes=constants_file.stat().st_size,
            last_modified=constants_file.stat().st_mtime,
            total_field_count=len(constants_dict)
        )
        
    except PermissionError as e:
        raise PermissionError(f"Cannot read constants file: {e}")
    except Exception as e:
        raise ValueError(f"Invalid constants file format: {e}")


def validate_constants_file(file_path: Path) -> bool:
    """
    Validate that constants file exists and has required structure.
    
    Args:
        file_path: Path to the constants Excel file
        
    Returns:
        True if file is valid, False otherwise
        
    Example:
        >>> is_valid = validate_constants_file(Path("constants.xlsx"))
        >>> if not is_valid: print("Invalid constants file")
    """
    # Check file exists and is readable
    if not file_path.exists():
        logger.warning(f"Constants file does not exist: {file_path}")
        return False
    
    if not file_path.is_file():
        logger.warning(f"Path is not a file: {file_path}")
        return False
        
    try:
        # Import openpyxl for validation
        import openpyxl
        
        # Try to load workbook (this will catch corrupted files)
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        # Validate "Pricing Setup" worksheet exists
        if "Pricing Setup" not in workbook.sheetnames:
            logger.warning(f"Required worksheet 'Pricing Setup' not found in {file_path}")
            logger.warning(f"Available worksheets: {workbook.sheetnames}")
            return False
        
        worksheet = workbook["Pricing Setup"]
        
        # Check Column C and D contain some data
        has_c_data = False
        has_d_data = False
        
        # Check first 50 rows for data
        for row in range(1, 51):
            if worksheet[f"C{row}"].value is not None:
                has_c_data = True
            if worksheet[f"D{row}"].value is not None:
                has_d_data = True
            
            # Early exit if both found
            if has_c_data and has_d_data:
                break
        
        if not has_c_data:
            logger.warning(f"Column C appears to be empty in {file_path}")
            return False
            
        if not has_d_data:
            logger.warning(f"Column D appears to be empty in {file_path}")
            return False
        
        logger.debug(f"Constants file validation successful: {file_path}")
        return True
        
    except ImportError:
        logger.error("openpyxl library not available - cannot validate Excel files")
        return False
    except Exception as e:
        logger.error(f"Error validating constants file {file_path}: {e}")
        return False


def parse_field_value_mapping(worksheet, col_c: str, col_e: str) -> Dict[str, str]:
    """
    Parse field name to value mapping from worksheet columns.
    
    Args:
        worksheet: Excel worksheet object
        col_c: Column containing field names (e.g., 'C')
        col_e: Column containing field values (e.g., 'E')
        
    Returns:
        Dictionary mapping field names to values
        
    Example:
        >>> mapping = parse_field_value_mapping(ws, 'C', 'E')
        >>> print(f"Found {len(mapping)} field mappings")
    """
    field_mapping = {}
    
    # Get maximum row to avoid infinite loop
    max_row = worksheet.max_row
    
    # Scan through all rows looking for field name/value pairs
    for row in range(1, max_row + 1):
        # Get field name from Column C
        field_cell = worksheet[f"{col_c}{row}"]
        value_cell = worksheet[f"{col_e}{row}"]
        
        # Skip empty rows or cells
        if field_cell.value is None or value_cell.value is None:
            continue
            
        # Convert to strings and trim whitespace
        field_name = str(field_cell.value).strip()
        field_value = str(value_cell.value).strip()
        
        # Skip empty strings after trimming
        if not field_name or not field_value:
            continue
            
        # Skip field values that are "None" or "Please Select" (placeholder values)
        if field_value.lower() in ['none', 'please select']:
            continue
            
        # Skip field names that are headers (end with colon) or formulas (start with =)
        if field_name.endswith(':') or field_name.startswith('='):
            continue
            
        # Store mapping
        field_mapping[field_name] = field_value
        logger.debug(f"Mapped field: '{field_name}' -> '{field_value}'")
    
    return field_mapping