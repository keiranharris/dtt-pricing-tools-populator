"""
Excel Data Populator Module

This module provides functions for writing matched data to target Excel spreadsheets.
Handles various cell types including dropdown selectors and preserves formatting.
Includes validation and error recovery capabilities.

Dependencies:
- dataclasses: Data structures
- pathlib: Path handling
- typing: Type hints
- openpyxl: Excel file operations

Author: DTT Pricing Tool Accelerator
Feature: 002-excel-data-population
"""

from dataclasses import dataclass
from pathlib import Path
from typing import List, Dict, Optional
import logging

# Import field matcher types
from field_matcher import FieldMatch, CellLocation

# Configure logging
logger = logging.getLogger(__name__)


@dataclass
class PopulationResult:
    """Results from populating data into Excel spreadsheet."""
    successful_fields: int
    failed_fields: int
    total_fields: int
    error_messages: List[str]
    populated_fields: List[str]
    
    @property
    def success_rate(self) -> float:
        """Calculate success rate as percentage."""
        if self.total_fields == 0:
            return 0.0
        return (self.successful_fields / self.total_fields) * 100
    
    def __str__(self) -> str:
        return f"Population Result: {self.successful_fields}/{self.total_fields} fields ({self.success_rate:.1f}%)"


def write_value_to_cell(worksheet, location: CellLocation, value: str) -> bool:
    """
    Write value to specific cell in Excel worksheet.
    
    Args:
        worksheet: Excel worksheet object
        location: Target cell location
        value: Value to write to the cell
        
    Returns:
        True if write was successful, False otherwise
        
    Example:
        >>> success = write_value_to_cell(ws, CellLocation(5, 2, "B5"), "Test Value")
        >>> if success: print("Value written successfully")
    """
    try:
        # Get target cell
        target_cell = worksheet.cell(row=location.row, column=location.column)
        
        # Store original value for validation
        original_value = target_cell.value
        
        # Write new value (this works for both regular cells and dropdowns)
        target_cell.value = value
        
        # Validate write was successful
        if target_cell.value == value:
            logger.debug(f"✅ Successfully wrote '{value}' to {location.cell_reference}")
            return True
        else:
            logger.warning(f"❌ Write validation failed at {location.cell_reference}: "
                         f"expected '{value}', got '{target_cell.value}'")
            return False
            
    except PermissionError as e:
        logger.error(f"Permission denied writing to {location.cell_reference}: {e}")
        return False
    except Exception as e:
        logger.error(f"Error writing to {location.cell_reference}: {e}")
        return False


def populate_matched_fields(target_file: Path, matches: List[FieldMatch]) -> PopulationResult:
    """
    Populate all matched fields in target Excel file.
    
    Args:
        target_file: Path to target Excel spreadsheet
        matches: List of field matches to populate
        
    Returns:
        Results summary of population operation
        
    Example:
        >>> result = populate_matched_fields(Path("output.xlsx"), field_matches)
        >>> print(f"Success rate: {result.success_rate:.1f}%")
    """
    if not matches:
        logger.warning("No field matches provided for population")
        return PopulationResult(
            successful_fields=0,
            failed_fields=0,
            total_fields=0,
            error_messages=["No field matches provided"],
            populated_fields=[]
        )
    
    try:
        # Import openpyxl for Excel operations
        import openpyxl
        
        # Open target Excel file
        logger.info(f"Opening target file for population: {target_file}")
        workbook = openpyxl.load_workbook(target_file)
        
        # Get "Pricing Setup" worksheet
        if "Pricing Setup" not in workbook.sheetnames:
            error_msg = f"Target worksheet 'Pricing Setup' not found in {target_file}"
            logger.error(error_msg)
            return PopulationResult(
                successful_fields=0,
                failed_fields=len(matches),
                total_fields=len(matches),
                error_messages=[error_msg],
                populated_fields=[]
            )
        
        worksheet = workbook["Pricing Setup"]
        
        # Track results
        successful_fields = 0
        failed_fields = 0
        error_messages = []
        populated_fields = []
        
        # Populate each matched field
        for match in matches:
            logger.info(f"Populating: {match.source_field} -> {match.target_location.cell_reference}")
            
            success = write_value_to_cell(worksheet, match.target_location, match.source_value)
            
            if success:
                successful_fields += 1
                populated_fields.append(f"{match.source_field} -> {match.target_location.cell_reference}")
            else:
                failed_fields += 1
                error_msg = f"Failed to populate {match.source_field} at {match.target_location.cell_reference}"
                error_messages.append(error_msg)
        
        # Save workbook
        logger.info("Saving populated Excel file...")
        workbook.save(target_file)
        logger.info(f"✅ Successfully saved populated file: {target_file}")
        
        # Create result summary
        result = PopulationResult(
            successful_fields=successful_fields,
            failed_fields=failed_fields,
            total_fields=len(matches),
            error_messages=error_messages,
            populated_fields=populated_fields
        )
        
        logger.info(f"Population complete: {result}")
        return result
        
    except ImportError:
        error_msg = "openpyxl library not available - cannot write Excel files"
        logger.error(error_msg)
        return PopulationResult(
            successful_fields=0,
            failed_fields=len(matches),
            total_fields=len(matches),
            error_messages=[error_msg],
            populated_fields=[]
        )
    except Exception as e:
        error_msg = f"Error during population: {e}"
        logger.error(error_msg)
        return PopulationResult(
            successful_fields=0,
            failed_fields=len(matches),
            total_fields=len(matches),
            error_messages=[error_msg],
            populated_fields=[]
        )


def validate_population_success(worksheet, matches: List[FieldMatch]) -> Dict[str, bool]:
    """
    Validate that all matched fields were populated correctly.
    
    Args:
        worksheet: Excel worksheet object
        matches: List of field matches to validate
        
    Returns:
        Dictionary mapping field names to validation success
        
    Example:
        >>> validation = validate_population_success(ws, matches)
        >>> failed_fields = [field for field, success in validation.items() if not success]
    """
    # Implementation will be added in Phase 3
    pass


def populate_matched_fields_xlwings(target_file: Path, matches: List[FieldMatch]) -> PopulationResult:
    """
    Populate matched fields using xlwings for .xlsb file support.
    
    This version uses xlwings to directly modify .xlsb files while preserving
    all original formatting, formulas, and functionality.
    
    Args:
        target_file: Path to the .xlsb target file
        matches: List of field matches to populate
        
    Returns:
        PopulationResult with success/failure statistics
        
    Example:
        >>> matches = [FieldMatch(...), FieldMatch(...)]
        >>> result = populate_matched_fields_xlwings(Path("template.xlsb"), matches)
        >>> print(f"Success rate: {result.success_rate:.1f}%")
        Success rate: 100.0%
    """
    import xlwings as xw
    
    logger.info(f"Opening target .xlsb file for population: {target_file}")
    
    successful_fields = 0
    failed_fields = 0 
    error_messages = []
    populated_fields = []
    
    try:
        with xw.App(visible=False) as app:
            wb = app.books.open(target_file)
            
            # Get Pricing Setup worksheet
            if 'Pricing Setup' not in [ws.name for ws in wb.sheets]:
                error_msg = "Pricing Setup worksheet not found in .xlsb file"
                logger.error(f"❌ {error_msg}")
                wb.close()
                return PopulationResult(
                    successful_fields=0,
                    failed_fields=len(matches),
                    total_fields=len(matches),
                    error_messages=[error_msg],
                    populated_fields=[]
                )
            
            ws = wb.sheets['Pricing Setup']
            
            # Populate each matched field
            for match in matches:
                try:
                    logger.info(f"Populating: {match.source_field} -> {match.target_location.cell_reference}")
                    logger.info(f"Writing value: '{match.source_value}' to cell {match.target_location.cell_reference}")
                    
                    # Adjust target cell: if we matched a label cell (E column), write to value cell (F column)
                    target_cell_ref = match.target_location.cell_reference
                    if target_cell_ref.startswith('E'):
                        # Convert E16 -> F16, E17 -> F17, etc.
                        target_cell_ref = 'F' + target_cell_ref[1:]
                        logger.info(f"Adjusting target cell from {match.target_location.cell_reference} to {target_cell_ref}")
                    
                    # Get the target cell and set the value
                    target_cell = ws.range(target_cell_ref)
                    
                    # Debug: Check current cell value
                    current_value = target_cell.value
                    logger.info(f"Current cell value: '{current_value}'")
                    
                    # Set the new value
                    target_cell.value = match.source_value
                    
                    # Verify the value was set
                    new_value = target_cell.value
                    logger.info(f"New cell value after write: '{new_value}'")
                    
                    if str(new_value) == str(match.source_value):
                        successful_fields += 1
                        populated_fields.append(match.source_field)
                        logger.info(f"✅ Successfully wrote '{match.source_value}' to {match.target_location.cell_reference}")
                    else:
                        logger.warning(f"❌ Value verification failed for {match.target_location.cell_reference}")
                        failed_fields += 1
                    
                except Exception as e:
                    error_msg = f"Failed to populate {match.source_field}: {e}"
                    logger.error(f"❌ {error_msg}")
                    error_messages.append(error_msg)
                    failed_fields += 1
            
            # Save the file (preserves .xlsb format)
            logger.info("Saving populated .xlsb file...")
            wb.save()
            logger.info(f"✅ Successfully saved populated file: {target_file}")
            
            wb.close()
            
    except Exception as e:
        error_msg = f"Failed to open/save .xlsb file: {e}"
        logger.error(f"❌ {error_msg}")
        error_messages.append(error_msg)
        failed_fields = len(matches)
        successful_fields = 0
    
    result = PopulationResult(
        successful_fields=successful_fields,
        failed_fields=failed_fields,
        total_fields=len(matches),
        error_messages=error_messages,
        populated_fields=populated_fields
    )
    
    logger.info(f"Population complete: {result}")
    return result


def populate_fields_in_worksheet(worksheet, matches: List[FieldMatch]) -> PopulationResult:
    """
    Populate matched fields using an already-open xlwings worksheet object.
    
    This function works with an existing xlwings worksheet object (part of an already-open workbook)
    instead of opening files. This enables consolidated Excel operations in a single session.
    
    Args:
        worksheet: xlwings worksheet object (already open)
        matches: List of field matches to populate
        
    Returns:
        PopulationResult with success/failure statistics
        
    Example:
        >>> with ExcelSessionManager(file_path) as session:
        ...     worksheet = session.get_worksheet("Pricing Setup")
        ...     result = populate_fields_in_worksheet(worksheet, matches)
        ...     print(f"Success rate: {result.success_rate:.1f}%")
    """
    logger.info(f"Populating {len(matches)} fields in existing worksheet session...")
    
    successful_fields = 0
    failed_fields = 0 
    error_messages = []
    populated_fields = []
    
    try:
        # Populate each matched field
        for match in matches:
            try:
                logger.info(f"Populating: {match.source_field} -> {match.target_location.cell_reference}")
                logger.info(f"Writing value: '{match.source_value}' to cell {match.target_location.cell_reference}")
                
                # Adjust target cell: if we matched a label cell (E column), write to value cell (F column)
                target_cell_ref = match.target_location.cell_reference
                if target_cell_ref.startswith('E'):
                    # Convert E16 -> F16, E17 -> F17, etc.
                    target_cell_ref = 'F' + target_cell_ref[1:]
                    logger.info(f"Adjusting target cell from {match.target_location.cell_reference} to {target_cell_ref}")
                
                # Get the target cell and set the value
                target_cell = worksheet.range(target_cell_ref)
                
                # Debug: Check current cell value
                current_value = target_cell.value
                logger.info(f"Current cell value: '{current_value}'")
                
                # Set the new value
                target_cell.value = match.source_value
                
                # Verify the value was set
                new_value = target_cell.value
                logger.info(f"New cell value after write: '{new_value}'")
                
                if str(new_value) == str(match.source_value):
                    successful_fields += 1
                    populated_fields.append(match.source_field)
                    logger.info(f"✅ Successfully wrote '{match.source_value}' to {target_cell_ref}")
                else:
                    logger.warning(f"❌ Value verification failed for {target_cell_ref}")
                    failed_fields += 1
                
            except Exception as e:
                error_msg = f"Failed to populate {match.source_field}: {e}"
                logger.error(f"❌ {error_msg}")
                error_messages.append(error_msg)
                failed_fields += 1
        
    except Exception as e:
        error_msg = f"Failed to populate worksheet: {e}"
        logger.error(f"❌ {error_msg}")
        error_messages.append(error_msg)
        failed_fields = len(matches)
        successful_fields = 0
    
    result = PopulationResult(
        successful_fields=successful_fields,
        failed_fields=failed_fields,
        total_fields=len(matches),
        error_messages=error_messages,
        populated_fields=populated_fields
    )
    
    logger.info(f"Worksheet population complete: {result}")
    return result