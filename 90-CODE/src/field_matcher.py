"""
Field Matcher Module

This module provides intelligent field matching capabilities for mapping between
constants files and target spreadsheets. Uses core content matching by stripping
decorative prefixes/suffixes and comparing middle content with fuzzy matching.

Dependencies:
- dataclasses: Data structures
- difflib: String similarity matching
- typing: Type hints
- re: Regular expressions

Author: DTT Pricing Tool Accelerator
Feature: 002-excel-data-population
"""

from dataclasses import dataclass
from typing import List, Dict, Optional
import logging
import difflib

# Configure logging
logger = logging.getLogger(__name__)


# Import SpecKit data models
from data_models import CellLocation, FieldMatch


def scan_worksheet_for_pricing_setup_fields(worksheet) -> List[CellLocation]:
    """
    Scan only the hard-coded columns for field names in 'Pricing Setup' sheet for performance.
    Uses global constants for column indices.
    """
    from constants import PRICING_SETUP_OUTPUT_FIELD_COL_IDX, PRICING_SETUP_OUTPUT_VALUE_COL_IDX
    cell_locations = []
    max_row = min(worksheet.max_row, 100)
    allowed_columns = [PRICING_SETUP_OUTPUT_FIELD_COL_IDX, PRICING_SETUP_OUTPUT_VALUE_COL_IDX]
    for row in range(1, max_row + 1):
        for col in allowed_columns:
            try:
                cell = worksheet.cell(row=row, column=col)
                if cell.value is None:
                    continue
                cell_text = str(cell.value).strip()
                if _is_potential_field_name(cell_text):
                    from openpyxl.utils import get_column_letter
                    col_letter = get_column_letter(col)
                    cell_ref = f"{col_letter}{row}"
                    location = CellLocation(row=row, column=col, cell_reference=cell_ref, content=cell_text)
                    cell_locations.append(location)
            except Exception as e:
                continue
    return cell_locations


def scan_worksheet_for_pricing_setup_fields_xlwings(worksheet) -> List[CellLocation]:
    """
    Scan only the hard-coded columns for field names in 'Pricing Setup' sheet for performance (xlwings version).
    Uses global constants for column indices.
    """
    from constants import PRICING_SETUP_OUTPUT_FIELD_COL_IDX, PRICING_SETUP_OUTPUT_VALUE_COL_IDX
    cell_locations = []
    max_row = 100  # Reasonable limit for performance
    allowed_columns = [PRICING_SETUP_OUTPUT_FIELD_COL_IDX, PRICING_SETUP_OUTPUT_VALUE_COL_IDX]  # E and F
    
    logger.info(f"DEBUG: Scanning xlwings worksheet columns {allowed_columns} (E,F) up to row {max_row}")
    
    for row in range(1, max_row + 1):
        for col in allowed_columns:
            try:
                # Convert column index to letter for xlwings
                col_letter = chr(64 + col)  # 5->E, 6->F
                cell_ref = f"{col_letter}{row}"
                cell_value = worksheet.range(cell_ref).value
                
                if cell_value and isinstance(cell_value, str):
                    cell_text = cell_value.strip()
                    if _is_potential_field_name(cell_text):
                        location = CellLocation(row=row, column=col, cell_reference=cell_ref, content=cell_text)
                        cell_locations.append(location)
                        if len(cell_locations) <= 10:  # Log first 10 findings
                            logger.info(f"DEBUG: Found potential field at {cell_ref}: '{cell_text}'")
            except Exception as e:
                continue
    
    logger.info(f"DEBUG: xlwings scanning found {len(cell_locations)} potential fields")
    return cell_locations


def core_string_match(source_field: str, target_field: str, strip_count: int = 2) -> float:
    """
    Compare core content of field names by stripping decorative characters.
    
    Args:
        source_field: Field name from constants file
        target_field: Field name from target spreadsheet
        strip_count: Number of characters to strip from each end
        
    Returns:
        Similarity score between 0.0 and 1.0
        
    Example:
        >>> score = core_string_match("01. Opportunity ID", "A. Opportunity ID:")
        >>> print(f"Match confidence: {score:.1%}")
    """
    # Handle empty or None inputs
    if not source_field or not target_field:
        return 0.0
    
    # Enhanced normalization for better matching
    source_normalized = normalize_field_name(source_field)
    target_normalized = normalize_field_name(target_field)
    
    # Handle empty fields after normalization
    if not source_normalized or not target_normalized:
        return 0.0
    
    # Use difflib for sequence matching
    similarity = difflib.SequenceMatcher(None, source_normalized, target_normalized).ratio()
    
    logger.debug(f"String match: '{source_field}' -> '{target_field}'")
    logger.debug(f"  Normalized: '{source_normalized}' -> '{target_normalized}' = {similarity:.2f}")
    
    return similarity


def normalize_field_name(field_name: str) -> str:
    """
    Normalize field names for better matching by removing common decorations and variations.
    
    Args:
        field_name: Original field name
        
    Returns:
        Normalized field name for comparison
    """
    import re
    
    if not field_name:
        return ""
    
    # Convert to lowercase
    normalized = field_name.lower().strip()
    
    # Remove asterisks (common in required fields)
    normalized = normalized.replace('*', '')
    
    # Remove common punctuation and decorations
    normalized = re.sub(r'[:\.\-\(\)\[\]]+$', '', normalized)  # Remove trailing punctuation
    normalized = re.sub(r'^[:\.\-\(\)\[\]]+', '', normalized)  # Remove leading punctuation
    
    # Remove numbers and bullets at start
    normalized = re.sub(r'^\d+[\.\)\-\s]*', '', normalized)
    
    # Remove common prefixes
    prefixes = ['a.', 'b.', 'c.', 'd.', 'e.', 'f.', 'g.', 'h.', 'i.', 'j.']
    for prefix in prefixes:
        if normalized.startswith(prefix):
            normalized = normalized[len(prefix):].strip()
            break
    
    # Normalize whitespace
    normalized = re.sub(r'\s+', ' ', normalized).strip()
    
    return normalized


def strip_decorations(field_name: str, strip_count: int = 2) -> str:
    """
    Remove decorative prefixes and suffixes from field names.
    
    Args:
        field_name: Original field name with potential decorations
        strip_count: Number of characters to remove from each end
        
    Returns:
        Core content of field name
        
    Example:
        >>> core = strip_decorations("01. Opportunity ID:")
        >>> print(f"Core content: '{core}'")
    """
    if not field_name or len(field_name) <= strip_count * 2:
        # For very short strings, return as-is
        return field_name.strip()
    
    # Strip specified number of characters from start and end
    core_content = field_name[strip_count:-strip_count]
    
    # Clean up any remaining whitespace
    return core_content.strip()


def scan_worksheet_for_fields(worksheet) -> List[CellLocation]:
    """
    Scan worksheet to find all cells that could contain field names.
    
    Args:
        worksheet: Excel worksheet object to scan
        
    Returns:
        List of cell locations containing potential field names
        
    Example:
        >>> locations = scan_worksheet_for_fields(worksheet)
        >>> print(f"Found {len(locations)} potential field locations")
    """
    cell_locations = []
    
    # Get worksheet dimensions
    max_row = min(worksheet.max_row, 100)  # Limit to first 100 rows for performance
    max_col = min(worksheet.max_column, 20)  # Limit to first 20 columns
    
    logger.debug(f"Scanning worksheet: {max_row} rows x {max_col} columns")
    
    # Scan all cells in the defined range
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            try:
                cell = worksheet.cell(row=row, column=col)
                
                # Skip empty cells
                if cell.value is None:
                    continue
                
                # Convert to string
                cell_text = str(cell.value).strip()
                
                # Filter out non-field-like content
                if _is_potential_field_name(cell_text):
                    # Convert column number to letter (1->A, 2->B, etc.)
                    from openpyxl.utils import get_column_letter
                    col_letter = get_column_letter(col)
                    cell_ref = f"{col_letter}{row}"
                    
                    location = CellLocation(
                        row=row,
                        column=col, 
                        cell_reference=cell_ref
                    )
                    
                    cell_locations.append(location)
                    logger.debug(f"Found potential field at {cell_ref}: '{cell_text[:30]}...'")
                    
            except Exception as e:
                logger.debug(f"Error scanning cell {row},{col}: {e}")
                continue
    
    logger.info(f"Found {len(cell_locations)} potential field locations")
    return cell_locations


def _is_potential_field_name(cell_text: str) -> bool:
    """
    Determine if a cell contains text that could be a field name.
    
    Args:
        cell_text: Text content of the cell
        
    Returns:
        True if text could be a field name, False otherwise
    """
    # Basic filtering rules
    if len(cell_text) < 3:  # Too short to be meaningful
        logger.debug(f"DEBUG: Filtered out (too short): '{cell_text}'")
        return False
    
    if len(cell_text) > 200:  # Too long to be a field name
        logger.debug(f"DEBUG: Filtered out (too long): '{cell_text[:50]}...'")
        return False
    
    # Skip pure numbers
    if cell_text.replace('.', '').replace('-', '').isdigit():
        logger.debug(f"DEBUG: Filtered out (pure number): '{cell_text}'")
        return False
    
    # Skip cells that look like formulas or references
    if cell_text.startswith('=') or cell_text.startswith('#'):
        logger.debug(f"DEBUG: Filtered out (formula/ref): '{cell_text}'")
        return False
    
    # Skip cells with only special characters
    if not any(c.isalpha() for c in cell_text):
        logger.debug(f"DEBUG: Filtered out (no alpha): '{cell_text}'")
        return False
    
    # Must contain some alphabetic characters (likely field names have words)
    alpha_count = sum(1 for c in cell_text if c.isalpha())
    if alpha_count < 3:  # Need at least 3 letters
        logger.debug(f"DEBUG: Filtered out (too few letters): '{cell_text}'")
        return False
    
    logger.debug(f"DEBUG: Accepted as potential field: '{cell_text}'")
    return True


def find_matching_fields(source_fields: Dict[str, str], target_sheet, threshold: float = 0.65) -> List[FieldMatch]:
    """
    Find best matches between source fields and target spreadsheet fields.
    
    Args:
        source_fields: Dictionary of field names to values from constants
        target_sheet: Target Excel worksheet to search
        threshold: Minimum similarity score for matches (default 0.8)
        
    Returns:
        List of high-confidence field matches
        
    Example:
        >>> matches = find_matching_fields(constants_data, target_worksheet)
        >>> print(f"Found {len(matches)} field matches")
    """
    if not source_fields:
        logger.info("No source fields provided for matching")
        return []
    
    # Use optimized scanning for Pricing Setup sheets
    logger.info("Scanning target worksheet for field locations...")
    try:
        # Check if this is a Pricing Setup worksheet by trying to access sheet name
        sheet_name = getattr(target_sheet, 'title', getattr(target_sheet, 'name', ''))
        if 'pricing setup' in sheet_name.lower():
            logger.info("Using optimized scanning for Pricing Setup sheet")
            target_locations = scan_worksheet_for_pricing_setup_fields(target_sheet)
        else:
            logger.info("Using generic scanning for worksheet")
            target_locations = scan_worksheet_for_fields(target_sheet)
    except Exception as e:
        logger.warning(f"Could not determine sheet type, using generic scanning: {e}")
        target_locations = scan_worksheet_for_fields(target_sheet)
    
    if not target_locations:
        logger.warning("No potential field locations found in target worksheet")
        return []
    
    matches = []
    
    # For each source field, find best matching target location
    for source_field, source_value in source_fields.items():
        best_match = None
        best_confidence = 0.0
        
        # Check against all target locations
        for location in target_locations:
            try:
                # Get target field text
                target_cell = target_sheet.cell(row=location.row, column=location.column)
                target_field = str(target_cell.value).strip()
                
                # Calculate similarity
                confidence = core_string_match(source_field, target_field)
                
                # Track best match for this source field
                if confidence > best_confidence and confidence >= threshold:
                    best_confidence = confidence
                    best_match = FieldMatch(
                        source_field=source_field,
                        target_location=location,
                        confidence=confidence,
                        source_value=source_value,
                        match_method="sequence_match"
                    )
                    
            except Exception as e:
                logger.debug(f"Error checking location {location}: {e}")
                continue
        
        # Add best match if found
        if best_match:
            matches.append(best_match)
            logger.info(f"✅ Match found: '{source_field}' -> {best_match.target_location.cell_reference} "
                       f"(confidence: {best_confidence:.1%})")
        else:
            logger.warning(f"❌ No match found for: '{source_field}' (threshold: {threshold:.1%})")
    
    logger.info(f"Field matching complete: {len(matches)}/{len(source_fields)} fields matched")
    return matches


def find_matching_fields_xlwings(source_fields: Dict[str, str], target_file, threshold: float = 0.65) -> List[FieldMatch]:
    """
    Find matching fields using xlwings for .xlsb file support.
    
    This is the xlwings version of find_matching_fields() that can handle .xlsb files
    directly without conversion, preserving all original formatting and functionality.
    
    Args:
        source_fields: Dictionary mapping field names to their values
        target_file: Path to the .xlsb target file
        threshold: Minimum confidence level for field matches (0.0-1.0)
        
    Returns:
        List of FieldMatch objects for successfully matched fields
        
    Example:
        >>> fields = {"Client Name": "Acme Corp", "Location": "Sydney"}
        >>> matches = find_matching_fields_xlwings(fields, Path("template.xlsb"))
        >>> print(f"Found {len(matches)} matches")
        Found 2 matches
    """
    import xlwings as xw
    from pathlib import Path
    
    logger.info("Scanning target .xlsb file for field locations using xlwings...")
    matches = []
    
    try:
        with xw.App(visible=False) as app:
            wb = app.books.open(target_file)
            
            # Check if Pricing Setup sheet exists
            sheet_names = [ws.name for ws in wb.sheets]
            if 'Pricing Setup' not in sheet_names:
                logger.error("❌ Target worksheet 'Pricing Setup' not found in .xlsb file")
                wb.close()
                return []
            
            ws = wb.sheets['Pricing Setup']
            
            # Scan a broader range for potential field locations
            # .xlsb files might have fields in different locations than .xlsx
            potential_locations = []
            
            # Debug: Log what we're actually finding
            debug_count = 0
            
            for row in range(1, 151):  # Scan first 150 rows
                for col in range(1, 51):  # Scan columns A-AX (more comprehensive)
                    try:
                        cell_value = ws.range(row, col).value
                        
                        # Debug logging for first few cells
                        if debug_count < 10 and cell_value is not None:
                            logger.debug(f"Debug: Found cell {ws.range(row, col).get_address(False, False)}: {repr(cell_value)}")
                            debug_count += 1
                        
                        if cell_value and isinstance(cell_value, str) and len(cell_value.strip()) > 3:
                            # Create cell reference (A1 notation)
                            cell_ref = ws.range(row, col).get_address(False, False)
                            potential_locations.append(CellLocation(row, col, cell_ref, cell_value.strip()))
                    except Exception as e:
                        # Skip problematic cells (e.g., merged cells, formulas, etc.)
                        continue
            
            logger.info(f"Found {len(potential_locations)} potential field locations")
            
            # Match fields using the same algorithm as the openpyxl version
            for source_field, source_value in source_fields.items():
                best_match = None
                best_confidence = 0.0
                
                for location in potential_locations:
                    try:
                        confidence = core_string_match(source_field, location.content)
                        
                        if confidence >= threshold and confidence > best_confidence:
                            best_confidence = confidence
                            best_match = FieldMatch(
                                source_field=source_field,
                                target_location=location,
                                confidence=confidence,
                                source_value=source_value,
                                match_method="core_string"
                            )
                            
                    except Exception as e:
                        logger.debug(f"Error checking location {location}: {e}")
                        continue
                
                # Add best match if found
                if best_match:
                    matches.append(best_match)
                    logger.info(f"✅ Match found: '{source_field}' -> {best_match.target_location.cell_reference} "
                               f"(confidence: {best_confidence:.1%}) [VALUE CELL]")
                else:
                    logger.warning(f"❌ No match found for: '{source_field}' (threshold: {threshold:.1%})")
            
            wb.close()
            
    except Exception as e:
        logger.error(f"❌ Error reading .xlsb file with xlwings: {e}")
        return []
    
    logger.info(f"Field matching complete: {len(matches)}/{len(source_fields)} fields matched")
    return matches


def find_matching_fields_in_worksheet(source_fields: Dict[str, str], worksheet, threshold: float = 0.65) -> List[FieldMatch]:
    """
    Find matching fields using an already-open xlwings worksheet object.
    
    This function works with an existing xlwings worksheet object (part of an already-open workbook)
    instead of opening files. This enables consolidated Excel operations in a single session.
    
    Args:
        source_fields: Dictionary mapping field names to their values
        worksheet: xlwings worksheet object (already open)
        threshold: Minimum confidence level for field matches (0.0-1.0)
        
    Returns:
        List of FieldMatch objects for successfully matched fields
        
    Example:
        >>> with ExcelSessionManager(file_path) as session:
        ...     worksheet = session.get_worksheet("Pricing Setup")
        ...     matches = find_matching_fields_in_worksheet(fields, worksheet)
        ...     print(f"Found {len(matches)} matches")
    """
    logger.info("Scanning worksheet for field locations using existing session...")
    matches = []
    
    try:
        # Use optimized scanning for Pricing Setup sheets
        try:
            # Check if this is a Pricing Setup worksheet by trying to access sheet name
            sheet_name = getattr(worksheet, 'name', '')
            if 'pricing setup' in sheet_name.lower():
                logger.info("Using optimized xlwings scanning for Pricing Setup sheet")
                potential_locations = scan_worksheet_for_pricing_setup_fields_xlwings(worksheet)
            else:
                logger.info("Using generic xlwings scanning for worksheet")
                # Fallback to original generic scanning for non-Pricing Setup sheets
                potential_locations = []
                for row in range(1, 51):  # Smaller fallback range
                    for col in range(1, 21):  # Smaller fallback range
                        try:
                            cell_ref = f"{chr(64+col)}{row}"
                            cell_value = worksheet.range(cell_ref).value
                            
                            if cell_value and isinstance(cell_value, str) and len(cell_value.strip()) > 3:
                                potential_locations.append(CellLocation(row, col, cell_ref, cell_value.strip()))
                        except Exception:
                            continue
        except Exception as e:
            logger.warning(f"Could not determine sheet type, using generic xlwings scanning: {e}")
            # Fallback to original generic scanning
            potential_locations = []
            for row in range(1, 51):  # Smaller fallback range
                for col in range(1, 21):  # Smaller fallback range
                    try:
                        cell_ref = f"{chr(64+col)}{row}"
                        cell_value = worksheet.range(cell_ref).value
                        
                        if cell_value and isinstance(cell_value, str) and len(cell_value.strip()) > 3:
                            potential_locations.append(CellLocation(row, col, cell_ref, cell_value.strip()))
                    except Exception:
                        continue
        
        logger.info(f"Found {len(potential_locations)} potential field locations")
        
        # Match fields using the same algorithm as the file-based version
        for source_field, source_value in source_fields.items():
            best_match = None
            best_confidence = 0.0
            
            for location in potential_locations:
                try:
                    confidence = core_string_match(source_field, location.content)
                    
                    if confidence >= threshold and confidence > best_confidence:
                        best_confidence = confidence
                        best_match = FieldMatch(
                            source_field=source_field,
                            target_location=location,
                            confidence=confidence,
                            source_value=source_value,
                            match_method="core_string_openpyxl"
                        )
                        
                except Exception as e:
                    logger.debug(f"Error checking location {location}: {e}")
                    continue
            
            # Add best match if found
            if best_match:
                matches.append(best_match)
                logger.info(f"✅ Match found: '{source_field}' -> {best_match.target_location.cell_reference} "
                           f"(confidence: {best_confidence:.1%}) [VALUE CELL]")
            else:
                logger.warning(f"❌ No match found for: '{source_field}' (threshold: {threshold:.1%})")
        
    except Exception as e:
        logger.error(f"❌ Error scanning worksheet: {e}")
        return []
    
    logger.info(f"Worksheet field matching complete: {len(matches)}/{len(source_fields)} fields matched")
    return matches